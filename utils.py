# modules/utils.py

import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl import load_workbook
import pandas as pd  # Add this import

def setup_logging(log_file='logs/trading_bot.log', level=logging.INFO):
    """Configure logging for the trading bot."""
    logging.basicConfig(
        level=level,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )

def center_align_and_autofit_excel(file_path):
    """
    Centers all content in all sheets of the Excel file both horizontally and vertically,
    and adjusts column widths to fit the content.
    """
    # Load the workbook
    workbook = load_workbook(file_path)

    # Iterate over all sheets
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]

        # Dictionary to track the maximum width needed for each column
        column_widths = {}

        # Iterate over all rows and columns
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                # Apply alignment to each cell
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Calculate the maximum width for each column
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    column_letter = get_column_letter(cell.column)
                    if column_letter not in column_widths:
                        column_widths[column_letter] = cell_length
                    else:
                        column_widths[column_letter] = max(column_widths[column_letter], cell_length)

        # Adjust the width of each column based on the maximum content length
        for column_letter, width in column_widths.items():
            ws.column_dimensions[column_letter].width = width + 2  # Add some padding for readability

    # Save the updated workbook
    workbook.save(file_path)
    logging.info(f"Content successfully centered, middle-aligned, and columns auto-fitted in all sheets of {file_path}")

def remove_timezone(df):
    """
    Removes timezone information from all datetime columns in the DataFrame.

    Parameters:
        df (pd.DataFrame): The DataFrame to process.

    Returns:
        pd.DataFrame: The processed DataFrame with timezone-naive datetime columns.
    """
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            # Check if the datetime column is timezone-aware
            if df[col].dt.tz is not None:
                df[col] = df[col].dt.tz_localize(None)
    return df

def assert_timezone_naive(df, df_name):
    """
    Asserts that all datetime columns in the DataFrame are timezone-naive.

    Parameters:
        df (pd.DataFrame): The DataFrame to check.
        df_name (str): Name of the DataFrame for logging purposes.

    Raises:
        ValueError: If any datetime column is timezone-aware.
    """
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            if df[col].dt.tz is not None:
                raise ValueError(f"Column '{col}' in DataFrame '{df_name}' is timezone-aware.")
