# modules/reporting.py

import os
import pandas as pd
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from modules.utils import center_align_and_autofit_excel, remove_timezone, assert_timezone_naive  # Updated import

def add_summary_header(ws, content, span_range, freeze_at, header_fill_color="FFD700", font_size=14):
    """
    Adds a stylized summary header to the given worksheet.

    Parameters:
        ws (Worksheet): The worksheet to modify.
        content (str): The text content of the header.
        span_range (str): The range of cells to merge for the header (e.g., 'A1:M3').
        freeze_at (str): The cell at which to freeze panes (e.g., 'A5').
        header_fill_color (str): Background color for the header (default: gold).
        font_size (int): Font size for the header text.
    """
    # Merge cells for the header
    ws.merge_cells(span_range)
    
    # Get the top-left cell of the merged range
    start_cell = ws[span_range.split(":")[0]]
    
    # Style the merged cells
    start_cell.value = content
    start_cell.font = Font(bold=True, size=font_size)
    start_cell.alignment = Alignment(horizontal="center", vertical="center")
    start_cell.fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
    
    # Freeze panes below the header
    ws.freeze_panes = freeze_at

def save_results(trades, missed_signals, daily_summary, final_summary, signals_data, stock_name, output_dir="."):
    """
    Save the trading results to an Excel file with stylized headers and formatting.

    Parameters:
        trades (pd.DataFrame): DataFrame containing all executed trades.
        missed_signals (pd.DataFrame): DataFrame containing all missed signals.
        daily_summary (pd.DataFrame): DataFrame containing daily capital summaries.
        final_summary (dict): Dictionary containing overall performance metrics.
        signals_data (pd.DataFrame): DataFrame containing generated signals.
        stock_name (str): Name of the stock.
        output_dir (str): Directory where the Excel file will be saved.
    """
    # Remove timezone information from all relevant DataFrames
    trades = remove_timezone(trades)
    signals_data = remove_timezone(signals_data)
    missed_signals = remove_timezone(missed_signals)
    daily_summary = remove_timezone(daily_summary)
    # Note: final_summary is a dict, so no datetime columns to process

    # Optional: Assert that all datetime columns are timezone-naive
    try:
        assert_timezone_naive(trades, "trades")
        assert_timezone_naive(signals_data, "signals_data")
        assert_timezone_naive(missed_signals, "missed_signals")
        assert_timezone_naive(daily_summary, "daily_summary")
    except ValueError as e:
        logging.critical(e)
        raise

    # Ensure the directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Append timestamp and stock name to the file name
    timestamp = pd.Timestamp.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_filename = os.path.join(output_dir, f"{stock_name}_{timestamp}.xlsx")

    # Rename the column Charges to Total Charges
    trades = trades.rename(columns={'Charges': 'Total Charges'})

    # Format prices and other numerical values to two decimal places
    trades = trades.round({
        'Price': 2,
        'Profit/Loss': 2,
        'Total Charges': 2,
        'Brokerage': 2,
        'STT': 2,
        'Transaction Charge': 2,
        'SEBI Charge': 2,
        'GST': 2,
        'Stamp Duty': 2
    })
    missed_signals = missed_signals.round({'Price': 2})
    daily_summary = daily_summary.round({'Start of Day Capital': 2, 'End of Day Capital': 2})

    # Prepare the Final Summary sheet
    final_summary_data = pd.DataFrame({
        'Metric': [
            'Initial Capital',
            'Final Capital',
            'Number of Unsold Stocks in Portfolio',
            'Leverage Used',
            'RSI_PERIOD',
            'MACD_FAST',
            'MACD_SLOW',
            'MACD_SIGNAL',
            'ATR_PERIOD',
            'ATR_THRESHOLD',
            'INTERVAL',
            'SYMBOL_TOKEN',
            'FROM_DATE',
            'TO_DATE',
            'TRADE_ALLOCATION',
            'TARGET_PROFIT_PERCENTAGE',
            'ATR_MULTIPLIER'
        ],
        'Value': [
            round(final_summary.get('Initial Capital', 0), 2),
            round(final_summary.get('Final Capital', 0), 2),
            final_summary.get('Current Stock Holding', 'N/A'),
            final_summary.get('Leverage Used', 'N/A'),
            final_summary.get('RSI_PERIOD', 'N/A'),
            final_summary.get('MACD_FAST', 'N/A'),
            final_summary.get('MACD_SLOW', 'N/A'),
            final_summary.get('MACD_SIGNAL', 'N/A'),
            final_summary.get('ATR_PERIOD', 'N/A'),
            final_summary.get('ATR_THRESHOLD', 'N/A'),
            final_summary.get('INTERVAL', 'N/A'),
            final_summary.get('SYMBOL_TOKEN', 'N/A'),
            final_summary.get('FROM_DATE', 'Unknown'),
            final_summary.get('TO_DATE', 'Unknown'),
            final_summary.get('TRADE_ALLOCATION', 'N/A'),
            final_summary.get('TARGET_PROFIT_PERCENTAGE', 'N/A'),
            final_summary.get('ATR_MULTIPLIER', 'N/A')
        ]
    })

    # Use Pandas ExcelWriter with openpyxl engine
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        # Create a new workbook and add sheets
        trades.to_excel(writer, sheet_name="Trades", index=False, startrow=4)
        add_summary_header(
            writer.book["Trades"],
            content=f"Trade Summary for {stock_name} from {final_summary.get('FROM_DATE', 'Unknown')} to {final_summary.get('TO_DATE', 'Unknown')}",
            span_range="A1:M3",
            freeze_at="A6"
        )

        # Style the column headers (row 5) in 'Trades' sheet
        ws_trades = writer.book["Trades"]
        blue_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")  # White font for contrast
        for cell in ws_trades[5]:  # Row 5 contains the column headers
            cell.fill = blue_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write other sheets
        sheets = {
            "Missed Signals": missed_signals,
            "Daily Summary": daily_summary,
            "Final Summary": final_summary_data,
            "Signals": signals_data
        }

        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.book[sheet_name]
            # Freeze the header row for other sheets
            ws.freeze_panes = ws["A2"]
            # Optionally, add headers styling similar to Trades
            if sheet_name != "Final Summary":  # Assuming Final Summary has different headers
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:  # First row contains the column headers
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Apply alignment and autofit
    center_align_and_autofit_excel(output_filename)
    logging.info(f"Results saved to {output_filename}")

def save_summary(summary_data, summary_file, from_date, to_date, buy_rsi_threshold, sell_rsi_threshold, trade_config, indicator_config, historical_data):
    """
    Save the performance summary to a styled Excel file with a summary and trade settings sheet.

    Parameters:
        summary_data (list of dict): List containing summary metrics for each stock.
        summary_file (str): Path where the summary Excel file will be saved.
        from_date (str): Start date of the trading period.
        to_date (str): End date of the trading period.
        buy_rsi_threshold (float): RSI threshold for buy signals.
        sell_rsi_threshold (float): RSI threshold for sell signals.
        trade_config (dict): Dictionary containing trade-related configurations.
        indicator_config (dict): Dictionary containing indicator-related configurations.
        historical_data (dict): Dictionary containing historical data configurations.
    """

    # Convert the summary data to a DataFrame
    summary_df = pd.DataFrame(summary_data)
    
    # Remove timezone information
    summary_df = remove_timezone(summary_df)
    
    # Optional: Assert that summary_df is timezone-naive
    try:
        assert_timezone_naive(summary_df, "summary_df")
    except ValueError as e:
        logging.critical(e)
        raise
    
    # Calculate the Total Net P/L
    total_net_pl = summary_df['Profit/Loss'].sum()
    
    # Add the "Net P/L (₹)" column with the total in the first row
    summary_df['Net P/L (₹)'] = ""
    summary_df.at[0, 'Net P/L (₹)'] = f"₹ {total_net_pl:,.2f}"
    
    # Add "Initial Capital For Each Stock" metric to the Trade Settings sheet only
    initial_capital = trade_config.get('INITIAL_CAPITAL')
    if initial_capital:
        initial_capital_formatted = f"₹ {initial_capital:,.2f}"
    else:
        initial_capital_formatted = "N/A"  # Fallback if not defined

    # Prepare the Trade Settings data with the new metric
    trade_settings = [
        ("Leverage Used", trade_config.get("LEVERAGE", "N/A")),
        ("RSI Period", indicator_config.get("RSI_PERIOD", "N/A")),
        ("MACD FAST", indicator_config.get("MACD_FAST", "N/A")),
        ("MACD SLOW", indicator_config.get("MACD_SLOW", "N/A")),
        ("MACD SIGNAL", indicator_config.get("MACD_SIGNAL", "N/A")),
        ("ATR Period", indicator_config.get("ATR_PERIOD", "N/A")),
        ("ATR Threshold", indicator_config.get("ATR_THRESHOLD", "N/A")),
        ("INTERVAL", historical_data.get("interval", "N/A")),
        ("TRADE ALLOCATION", f"{trade_config.get('TRADE_ALLOCATION', 0) * 100}%"),
        ("TARGET PROFIT (%)", f"{trade_config.get('TARGET_PROFIT_PERCENTAGE', 0)}%"),
        ("ATR Multiplier", trade_config.get("ATR_MULTIPLIER", "N/A")),
        ("Buy RSI Setting", f">{buy_rsi_threshold}"),
        ("Sell RSI Setting", f">{sell_rsi_threshold}"),
        ("Initial Capital For Each Stock (₹)", initial_capital_formatted)
    ]
    
    trade_settings_df = pd.DataFrame(trade_settings, columns=["Metrics", "Value"])

    # Create a new Excel workbook using Pandas ExcelWriter
    with pd.ExcelWriter(summary_file, engine='openpyxl') as writer:
        # Write Summary sheet
        summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=4)
        ws_summary = writer.book["Summary"]
        header_text = f"Performance Summary from {from_date} to {to_date}"
        add_summary_header(ws_summary, content=header_text, span_range="A1:D3", freeze_at="A6")

        # Style the column headers (row 5) in 'Summary' sheet
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")  # White font for contrast
        for cell in ws_summary[5]:  # Row 5 contains the column headers
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Style the "Net P/L (₹)" cell in the first row
        net_pl_cell = ws_summary["D6"]  # Adjust if startrow is different
        net_pl_cell.font = Font(bold=True)
        net_pl_cell.alignment = Alignment(horizontal="center", vertical="center")
        net_pl_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Write Trade Settings sheet
        trade_settings_df.to_excel(writer, sheet_name="Trade Settings", index=False, startrow=4)
        ws_settings = writer.book["Trade Settings"]
        header_text = "Trade Settings Overview"
        add_summary_header(ws_settings, content=header_text, span_range="A1:B3", freeze_at="A5")

        # Style the column headers (row 5) in 'Trade Settings' sheet
        for cell in ws_settings[5]:  # Row 5 contains the column headers
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply alignment and autofit
    center_align_and_autofit_excel(summary_file)
    logging.info(f"Styled performance summary saved to {summary_file}")
